#!/usr/bin/env python3
"""Generate backend/app/utils/catalog_data.py from dataset/products.xlsx."""
import openpyxl

WB = "dataset/products.xlsx"

CATEGORY_META = {
    "Leaves":        {"slug": "leaves",        "display_order": 9,  "emoji": "🌿", "unit": "bunch", "diet": "veg"},
    "Pickle Items":  {"slug": "pickle-items",  "display_order": 10, "emoji": "🫙", "unit": "jar",   "diet": "veg"},
    "Karam Powders": {"slug": "karam-powders", "display_order": 8,  "emoji": "🧂", "unit": "pack",  "diet": "veg"},
    "Grocery":       {"slug": "grocery",       "display_order": 7,  "emoji": "🛒", "unit": "pack",  "diet": "veg"},
    "Spices":        {"slug": "spices",        "display_order": 5,  "emoji": "🌶️", "unit": "pack",  "diet": "veg"},
}

# per-product (emoji, unit) overrides for dataset items
PRODUCT_META = {
    "Mehandi Leaves (Fresh)":    ("🌿", "bunch"),
    "Betel Leaves":              ("🌿", "bunch"),
    "Mango Leaves":              ("🌿", "bunch"),
    "Neem Leaves":               ("🌿", "bunch"),
    "Bhel Patra":                ("🌿", "bunch"),
    "Anand Finger Millets":      ("🌾", "pack"),
    "Anand Little Millet":       ("🌾", "pack"),
    "Premium Small Peanuts":     ("🥜", "pack"),
    "Deep Sooji Rava":           ("🌾", "pack"),
    "Bambino Roasted Vermicelli":("🍝", "pack"),
    "Turmeric Powder":           ("🟡", "200g"),
}

CURATED_CATEGORIES = [
    {"category": "Vegetables", "slug": "vegetables", "display_order": 1},
    {"category": "Fruits",     "slug": "fruits",     "display_order": 2},
    {"category": "Dairy",      "slug": "dairy",      "display_order": 3},
    {"category": "Grains",     "slug": "grains",     "display_order": 4},
    {"category": "Spices",     "slug": "spices",     "display_order": 5},
    {"category": "Snacks",     "slug": "snacks",     "display_order": 6},
]

# (name, category, emoji, price, unit, diet, description) — curated Store2Home catalog.
# "Turmeric Powder" duplicates a dataset product; curated values win, search terms still apply.
CURATED_PRODUCTS = [
    ("Fresh Tomatoes",     "Vegetables", "🍅", 2.49, "lb",    "veg",
     "Ripe, farm-fresh tomatoes — perfect for curries, salads and chutneys."),
    ("Onions",             "Vegetables", "🧅", 1.29, "lb",    "veg",
     "Crisp red onions, ideal for tadka and daily cooking."),
    ("Potatoes",           "Vegetables", "🥔", 1.99, "lb",    "veg",
     "Sturdy all-purpose potatoes for curries, fries and roasts."),
    ("Cilantro",           "Vegetables", "🌿", 0.99, "bunch", "veg",
     "Fresh coriander leaves to finish any dish."),
    ("Green Chili",        "Vegetables", "🫑", 1.49, "lb",    "veg",
     "Hot green chillies, great for curries and pickles."),
    ("Cucumber",           "Vegetables", "🥒", 1.19, "lb",    "veg",
     "Cool, crisp cucumbers for salads and raita."),
    ("Bananas",            "Fruits",     "🍌", 0.59, "lb",    "veg",
     "Sweet ripe bananas — a household staple."),
    ("Apples (Gala)",      "Fruits",     "🍎", 2.99, "lb",    "veg",
     "Crunchy, sweet Gala apples."),
    ("Mangoes (Alphonso)", "Fruits",     "🥭", 3.49, "each",  "veg",
     "Premium Alphonso mangoes in season."),
    ("Whole Milk",         "Dairy",      "🥛", 3.99, "gal",   "veg",
     "Fresh whole milk, delivered chilled."),
    ("Curd / Yogurt",      "Dairy",      "🥣", 2.29, "qt",    "veg",
     "Thick, fresh-set yogurt for everyday meals."),
    ("Paneer",             "Dairy",      "🧀", 4.99, "lb",    "veg",
     "Soft Indian cottage cheese, ready to cook."),
    ("Basmati Rice",       "Grains",     "🍚", 8.99, "5lb",   "veg",
     "Aged long-grain basmati rice, fluffy every time."),
    ("Toor Dal",           "Grains",     "🥜", 4.49, "lb",    "veg",
     "Split pigeon peas — the base for a classic dal."),
    ("Wheat Atta",         "Grains",     "🌾", 6.99, "5lb",   "veg",
     "Whole-wheat flour for soft rotis and parathas."),
    ("Turmeric Powder",    "Spices",     "🟡", 2.79, "200g",  "veg",
     "Pure ground turmeric for cooking and wellness."),
    ("Red Chili Powder",   "Spices",     "🌶️", 2.99, "200g",  "veg",
     "Fiery red chilli powder for authentic heat."),
    ("Cumin Seeds (Jeera)", "Spices",    "⚫", 3.49, "200g",  "veg",
     "Aromatic cumin seeds for tadka and masalas."),
    ("Chicken (Leg Quarters)", "Snacks", "🍗", 2.99, "lb",    "nonveg",
     "Fresh chicken leg quarters, cleaned and ready."),
    ("Eggs (Dozen)",       "Snacks",     "🥚", 3.49, "12",    "nonveg",
     "Farm-fresh eggs, one dozen."),
    ("Murukulu",           "Snacks",     "🥨", 5.99, "pack",  "veg",
     "Crispy Andhra-style murukulu snack."),
    ("Biscuits (Parle-G)", "Snacks",     "🍪", 1.99, "pack",  "veg",
     "Classic Parle-G biscuits — a family favorite."),
]

# regional aliases for the curated (non-dataset) products
CURATED_TERMS = {
    "Fresh Tomatoes":      [("tomato", "alias", "English"), ("tamatar", "regional", "Hindi"), ("thakkali", "regional", "Tamil"), ("tamata", "regional", "Telugu")],
    "Onions":              [("onion", "alias", "English"), ("pyaz", "regional", "Hindi"), ("ulli", "regional", "Tamil"), ("ullipaya", "regional", "Telugu")],
    "Potatoes":            [("potato", "alias", "English"), ("aalu", "regional", "Hindi"), ("urulaikizhangu", "regional", "Tamil"), ("bangaladumpa", "regional", "Telugu")],
    "Cilantro":            [("coriander", "alias", "English"), ("dhania", "regional", "Hindi"), ("kothamalli", "regional", "Tamil"), ("kothimeera", "regional", "Telugu")],
    "Green Chili":         [("green chilli", "alias", "English"), ("hari mirch", "regional", "Hindi"), ("pachai milagai", "regional", "Tamil"), ("pacchi mirchi", "regional", "Telugu")],
    "Cucumber":            [("cucumber", "alias", "English"), ("khira", "regional", "Hindi"), ("vellari", "regional", "Tamil"), ("dosakaya", "regional", "Telugu")],
    "Bananas":             [("banana", "alias", "English"), ("kela", "regional", "Hindi"), ("valaipazham", "regional", "Tamil"), ("arati", "regional", "Telugu")],
    "Apples (Gala)":       [("apple", "alias", "English"), ("seb", "regional", "Hindi"), ("apple pazham", "regional", "Tamil"), ("appil", "regional", "Telugu")],
    "Mangoes (Alphonso)":  [("mango", "alias", "English"), ("aam", "regional", "Hindi"), ("mango pazham", "regional", "Tamil"), ("mamidi", "regional", "Telugu")],
    "Whole Milk":          [("milk", "alias", "English"), ("doodh", "regional", "Hindi"), ("paal", "regional", "Tamil"), ("paalu", "regional", "Telugu")],
    "Curd / Yogurt":       [("curd", "alias", "English"), ("yogurt", "alias", "English"), ("dahi", "regional", "Hindi"), ("thayir", "regional", "Tamil"), ("perugu", "regional", "Telugu")],
    "Paneer":              [("paneer", "alias", "English"), ("chena", "regional", "Hindi")],
    "Basmati Rice":        [("basmati", "alias", "English"), ("biryani rice", "alias", "English"), ("basmati annam", "regional", "Telugu")],
    "Toor Dal":            [("toor dal", "alias", "English"), ("arhar dal", "regional", "Hindi"), ("thuvaram paruppu", "regional", "Tamil"), ("kandi pappu", "regional", "Telugu")],
    "Wheat Atta":          [("atta", "alias", "English"), ("wheat flour", "alias", "English"), ("gehu ka atta", "regional", "Hindi"), ("godhuma pindi", "regional", "Telugu")],
    "Red Chili Powder":    [("red chilli powder", "alias", "English"), ("lal mirch", "regional", "Hindi"), ("kandam", "regional", "Telugu")],
    "Cumin Seeds (Jeera)": [("jeera", "alias", "English"), ("cumin", "alias", "English"), ("zeera", "regional", "Hindi"), ("seeragam", "regional", "Tamil"), ("jilakara", "regional", "Telugu")],
    "Chicken (Leg Quarters)": [("chicken", "alias", "English"), ("murgi", "regional", "Hindi"), ("kozi", "regional", "Tamil"), ("kodi", "regional", "Telugu")],
    "Eggs (Dozen)":        [("eggs", "alias", "English"), ("ande", "regional", "Hindi"), ("mutta", "regional", "Tamil"), ("guddu", "regional", "Telugu")],
    "Murukulu":            [("murukku", "alias", "English"), ("chakli", "alias", "English"), ("murukulu", "alias", "Telugu")],
    "Biscuits (Parle-G)":  [("parle-g", "alias", "English"), ("biscuit", "alias", "English")],
}


def esc(v):
    return repr(str(v))


def main():
    wb = openpyxl.load_workbook(WB, data_only=True)
    products = [r for r in wb["products"].iter_rows(min_row=2, values_only=True) if r and r[0]]
    terms = [r for r in wb["search_terms"].iter_rows(min_row=2, values_only=True) if r and r[0]]

    lines = []
    lines.append('"""Auto-generated catalog data (from dataset/products.xlsx + curated catalog).')
    lines.append('Do not edit by hand — regenerate with dataset/generate_catalog_data.py.')
    lines.append('"""')
    lines.append("")

    # categories (curated + dataset)
    lines.append("CATALOG = [")
    curated_cat_names = {c["category"] for c in CURATED_CATEGORIES}
    for cat in CURATED_CATEGORIES:
        lines.append(f'    {{"category": {esc(cat["category"])}, "slug": {esc(cat["slug"])}, "display_order": {cat["display_order"]}}},')
    for name, meta in CATEGORY_META.items():
        if name in curated_cat_names:
            continue
        lines.append(f'    {{"category": {esc(name)}, "slug": {esc(meta["slug"])}, "display_order": {meta["display_order"]}}},')
    lines.append("]")
    lines.append("")

    # merged products (curated win over dataset on name collision)
    dataset_by_name = {p[0]: p for p in products}
    names = set(dataset_by_name) | set(p[0] for p in CURATED_PRODUCTS)
    curated_by_name = {p[0]: p for p in CURATED_PRODUCTS}

    lines.append("PRODUCTS = [")
    for name in sorted(names):
        if name in curated_by_name:
            lines.append(f'    {curated_by_name[name]!r},')
        else:
            p = dataset_by_name[name]
            pad = p + (None,) * 8
            category, desc, price = pad[1], pad[3], pad[4]
            emoji, unit = PRODUCT_META.get(name, (CATEGORY_META.get(category, {}).get("emoji", "🛒"),
                                                  CATEGORY_META.get(category, {}).get("unit", "each")))
            diet = "veg"
            lines.append(f'    ({esc(name)}, {esc(category)}, {esc(emoji)}, {price}, {esc(unit)}, {esc(diet)}, {esc(desc)}),')
    lines.append("]")
    lines.append("")

    # search terms (dataset)
    lines.append("SEARCH_TERMS = [")
    for t in terms:
        pname, term, term_type, language = (list(t) + [None] * 4)[:4]
        lines.append(f'    ({esc(pname)}, {esc(term)}, {esc(term_type)}, {esc(language)}),')
    # curated terms
    for pname, tlist in CURATED_TERMS.items():
        for term, term_type, language in tlist:
            lines.append(f'    ({esc(pname)}, {esc(term)}, {esc(term_type)}, {esc(language)}),')
    lines.append("]")
    lines.append("")

    out = "backend/app/utils/catalog_data.py"
    with open(out, "w") as f:
        f.write("\n".join(lines))
    n_terms = len(terms) + sum(len(v) for v in CURATED_TERMS.values())
    print(f"wrote {out}: {len(names)} products, {n_terms} search terms")


if __name__ == "__main__":
    main()
